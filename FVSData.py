import csv
from openpyxl import Workbook
from openpyxl import load_workbook

def turnToList(list1): #puts all the columns they want into a list
    endIndex=list1.find(',')
    column_ans=[]
    while(endIndex!=-1):
        word=list1[0:endIndex]
        word=word.strip()
        print word
        word=word.lower()
        print word
        column_ans.append(word.lower())
        list1=list1[endIndex+1:]
        endIndex=list1.find(',')
    column_ans.append(list1.strip().lower())
    return column_ans

def findWords(list1): #puts all the columns they want into a list
    endIndex=list1.find(',')
    column_ans=[]
    while(endIndex!=-1):#basic code. Takes the input as a sentence, or a list of values separated by commas, and then puts them in a python lost
        word=list1[0:endIndex]
        column_ans.append(word.strip())
        list1=list1[endIndex+1:]
        endIndex=list1.find(',')
    column_ans.append(list1.strip())
    return column_ans

def getAllColumns(ws):  # returns a list of a list of columns in each tab
    colList = []
    tabNameList = ws.get_sheet_names()
    sheetList = []
    for name in tabNameList:
        sheetList.append(ws.get_sheet_by_name(name))
    for tab in sheetList:
        colList.append(getColumns(tab))  # list of lists, turn to single list later
    return colList

def getColumns(list): #returns a list of the columns in ONE tab
    columns = []
    for col in list.columns:
        if(col==None):
            break
        else:
            columns.append(col[0].value)
    return columns

def hashify(list, wb):
    dict={ }
    i =-1
    for a in list:
        i=i+1
        d=1
        for b in a:
            dict[b.lower()]= {'col':d, 'sheet': wb[i]}
            d=d+1
    return dict

def maxList(list):
    ans=0
    for val in list:
        if len(val)>ans:
            ans=len(val)
    return ans

def printAllColumns(column):
    for a in column:
        for b in a:
            print b

def writeCsvAsCsv(ans,f): #writes a csv file as a csv
    fileName = raw_input("What would you like to name this file? (please include the .csv)")
    with open(fileName, 'wb')as fp:  # change the data.csv to the name they put in .csv
        a = csv.writer(fp)
        a.writerows(ans)
    f.close()

def writeCsvAsXlsx(ans): #writes a csv file as an xlsx
    wbfinal = Workbook()
    wsfinal = wbfinal.get_sheet_by_name("Sheet")

    for counter1, item1 in enumerate(ans):
        for counter2, item2 in enumerate(ans[counter1]):
            c = wsfinal.cell(row=counter1 + 1, column=counter2 + 1)
            c.value = ans[counter1][counter2]
    saveFile = raw_input("What would you like to save this file as? Please remember to put the .xlsx ")  # change so that it automatically sets the file type
    wbfinal.save(saveFile)

def saveXlsxAsXlsx(wantedColumns,wb,dict): #saves an xlsx file as an xlsx
    wbfinal=Workbook()
    wsfinal=wbfinal.get_sheet_by_name("Sheet")
    counter=1
    for a in wantedColumns:#takes the column number of the particular column A being searched from the hash table
        tmpcol=dict[a]['col']
       # print tmpcol
        tmpsheet=dict[a]['sheet'] #takes the sheet name again, from the hash table
        #print tmpsheet
        for i in range(1,len(wb.get_sheet_by_name(tmpsheet).rows)+1): #copies the values from the input sheet to the output sheet
            c=wsfinal.cell(row=i, column=counter)
            c.value = wb.get_sheet_by_name(tmpsheet).rows[i - 1][tmpcol-1].value
        counter=counter+1
    saveFile=raw_input("What would you like to save this file as? Please remember to put the .xlsx ")
    wbfinal.save(saveFile)

def saveXlsxAsCsv(wanted,wb,dict): #saves an xlsx file as a csv
    wbfinal = Workbook()
    wsfinal = wbfinal.get_sheet_by_name("Sheet")#when making a new workbook, the default sheet name is "Sheet"
    counter = 1
    ans = []
    for a in wanted:
        tmp=[]
        tmpcol = dict[a]['col']
        tmpsheet = dict[a]['sheet']
        for i in range(1, len(wb.get_sheet_by_name(tmpsheet).rows) + 1):
            c = wb.get_sheet_by_name(tmpsheet).rows[i - 1][tmpcol - 1].value
            tmp.append(c)
        counter = counter + 1
        ans.append(tmp)

    tmp=[]
    final=[]
    for i in range(1, maxList(ans)+1):
        tmp = []
        for val in ans:
            if i<=len(val):
                tmp.append(val[i-1])
            else:
                tmp.append("None")
        final.append(tmp)

   # print final

    fileName = raw_input("What would you like to name this file? (please include the .csv)")
    with open(fileName, 'wb')as fp:  # change the data.csv to the name they put in .csv
        a = csv.writer(fp)
        a.writerows(final)


class fileInput(object):

    def value(self):
        return self.Val

    def __init__(self):
        file = raw_input("Please enter the name of the file: ")
        self.File=file
        if self.File[-3:].lower() == "csv":
            self.Val=0
        else:
            self.Val=1


    def initialize(self):
        if self.Val==0:
            self.F = open(self.File)
            self.csv_f = csv.reader(self.F)
            self.Columns = []
            for row in self.csv_f:
                self.Columns = row
                break

            self.F.seek(0)
          #  print "Here are the found columns! "
          #  print columns
            self.Dictionary = {}
            i = 0
            for a in self.Columns:
                self.Dictionary[a.lower()] = i
                i = i + 1
        else:
            self.wb = Workbook()

            self.wb = load_workbook(self.File)

            self.listCol = getAllColumns(self.wb)

            self.dict = hashify(self.listCol, self.wb.get_sheet_names())

    def printCol(self):
        if(self.Val==0):
            print "Here are the found columns! "
            print self.Columns
        else:
            print "Here are the found columns! "
            printAllColumns(self.listCol)

    def save(self):
        saveAs = raw_input("Would you like this to save as a csv or an xlsx? ")
        if(saveAs.lower()=="csv"):
            if(self.Val==0):
                writeCsvAsCsv(self.ans, self.F)
            else:
                saveXlsxAsCsv(self.wantedColumns,self.wb,self.dict)
        elif(saveAs.lower()=="xlsx"):
            if(self.Val==0):
                writeCsvAsXlsx(self.ans)
            else:
                saveXlsxAsXlsx(self.wantedColumns,self.wb,self.dict)

        else:
            print("Sorry, I couldn't understand that. Please type csv or xlsx")


    def getCols(self):
        if(self.Val==0):
            if (raw_input("Would you like the default output, or would you like to choose your own columns? Type 0 for default or 1 to choose your own ") == '0'):
                self.stuff = ['standid', 'plot_id', 'tree_id', 'species', 'tpa', 'dbh', 'dg', 'site_species', 'site_index']
            else:
                columnName = raw_input(
                    "Which columns would you like to extract? Please separate the column names by commas. ")

                self.stuff = findWords(columnName)
            self.indices = []

            # finds indices of the columns needed


            for whatsNeeded in self.stuff:
                print whatsNeeded
                tmp = self.Dictionary[whatsNeeded.lower()]
              #  print tmp
                self.indices.append(tmp)
              #  print indices

            self.ans = []

            for row in self.csv_f:
                tmp = []
                for val in self.indices:
                    tmp.append(row[val])
                self.ans.append(tmp)

        else:
            if (raw_input("Would you like the default output, or would you like to choose your own columns? Type 0 for default or 1 to choose your own ") == '0'):
                self.wantedColumns = ['standid', 'plot_id', 'tree_id', 'species', 'tpa', 'dbh', 'dg', 'site_species','site_index']
            else:

                self.wantedColumns = raw_input("Which columns would you like? ")
                self.wantedColumns = turnToList(self.wantedColumns)

